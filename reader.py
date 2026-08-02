# -*- coding: utf-8 -*-
"""读取 Excel / CSV 文件，并把每一格的内容解析成"秒"。"""

import os
import re
import csv
import datetime

EXT_KIND = {
    ".xlsx": "xlsx",
    ".xlsm": "xlsx",
    ".xls": "xls",
    ".csv": "csv",
}


class CsvData:
    def __init__(self, rows, encoding):
        self.rows = rows
        self.encoding = encoding


def _detect_encoding(path):
    with open(path, "rb") as f:
        raw = f.read()
    for enc in ("utf-8-sig", "utf-8", "gbk"):
        try:
            raw.decode(enc)
            return enc
        except UnicodeDecodeError:
            continue
    return "utf-8"


def load_file(path):
    ext = os.path.splitext(path)[1].lower()
    kind = EXT_KIND.get(ext)
    if kind is None:
        raise ValueError("不支持的格式：%s（仅支持 .xlsx / .xls / .csv）" % ext)

    if kind == "xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(path)
        return wb, wb.sheetnames, kind

    if kind == "xls":
        import xlrd
        try:
            wb = xlrd.open_workbook(path, formatting_info=True)
        except Exception:
            wb = xlrd.open_workbook(path)
        return wb, wb.sheet_names(), kind

    # csv
    encoding = _detect_encoding(path)
    with open(path, "r", encoding=encoding, newline="") as f:
        rows = list(csv.reader(f))
    return CsvData(rows, encoding), ["数据"], kind


def _is_time_cell(value, fmt):
    if isinstance(value, (datetime.time, datetime.datetime, datetime.timedelta)):
        return True
    f = (fmt or "").lower()
    return ("h" in f) or ("[m]" in f) or ("[s]" in f)


def read_all_columns(handle, kind, sheet_name):
    """读取工作表的所有列，返回 {"A": [...], "B": [...], ...}。

    只返回有非空值的列，每列是按行排列的单元格值列表。
    """
    if kind == "xlsx":
        ws = handle[sheet_name]
        cols = {}
        for c in range(1, ws.max_column + 1):
            letter = get_column_letter(c)
            vals = []
            for r in range(1, ws.max_row + 1):
                cell = ws.cell(r, c)
                v = cell.value
                if v is None:
                    continue
                if isinstance(v, str) and not v.strip():
                    continue
                vals.append((v, _is_time_cell(v, cell.number_format)))
            if vals:
                cols[letter] = vals
        return cols

    if kind == "xls":
        import xlrd
        sh = handle.sheet_by_name(sheet_name)
        cols = {}
        for c in range(sh.ncols):
            letter = get_column_letter(c + 1)
            vals = []
            for r in range(sh.nrows):
                cell = sh.cell(r, c)
                v = cell.value
                if v is None or v == "":
                    continue
                vals.append((v, cell.ctype == xlrd.XL_CELL_DATE))
            if vals:
                cols[letter] = vals
        return cols

    # csv：只有一列
    return {"A": read_column_a(handle, kind, sheet_name)}


def get_column_letter(n):
    """1 -> A, 2 -> B, ... 27 -> AA"""
    letter = ""
    while n:
        n, rem = divmod(n - 1, 26)
        letter = chr(65 + rem) + letter
    return letter


_FULLWIDTH = {ord("："): ":", ord("．"): ".", ord("，"): ",", ord("　"): " "}
_FULLWIDTH.update({ord("０") + i: str(i) for i in range(10)})


def _normalize_symbols(s):
    """全角符号/数字转半角，避免中英文混用导致解析失败。"""
    return s.translate(_FULLWIDTH)


class _TimeInvalid(object):
    """哨兵：识别为时间格式但非法（如秒数 >=60）。"""
    pass


_TIME_INVALID = _TimeInvalid()


def _parse_time_text(s):
    """把多种写法的 分:秒 文本解析成秒。

    返回：
      - 秒数（int）：解析成功
      - None：不是时间格式（可尝试纯数字解析）
      - _TIME_INVALID：是时间格式但非法（如秒数 >=60），应判整行无效
    """
    # 纯秒数 + 单位：120s / 120秒 / 120 S / 120 秒
    m = re.match(r"^\s*(\d+)\s*[s秒]\s*$", s, re.IGNORECASE)
    if m:
        return int(m.group(1))

    # 文字写法：2分27秒 / 2m27s / 2'27"，必须含标记字符才按此解析
    #（否则纯整数 90 会被误拆成 9分0秒）
    if re.search(r"[分秒m\u2032\u2033'\"`]", s):
        m = re.match(r"^(\d+)\s*[分m\u2032'\u0060]\s*(\d{1,2})\s*[秒s\u2033\"]?$", s)
        if m:
            minutes, seconds = int(m.group(1)), int(m.group(2))
            if seconds < 60:
                return minutes * 60 + seconds
            return _TIME_INVALID

    # 数字分隔写法：2:27 / 2.27 / 2,27
    parts = re.split(r"[:：.．,，]", s)
    if len(parts) >= 2:
        # 用 . 或 ，分隔的多段（如 1.2.3 / 12,34,56）属于奇怪数据，
        # 分.秒最多 2 段，只有 : 分隔才允许 3 段（时:分:秒）
        dot_parts = re.split(r"[.．,，]", s)
        if len(dot_parts) > 2 and ":" not in s and "：" not in s:
            return _TIME_INVALID
        try:
            nums = [int(float(p)) for p in parts]
        except ValueError:
            return None
        if len(nums) == 2:
            m, sec = nums
            if sec < 60:
                return m * 60 + sec
            return _TIME_INVALID
        if len(nums) == 3:
            h, m, sec = nums
            if sec < 60 and m < 60:
                return h * 3600 + m * 60 + sec
            return _TIME_INVALID
    return None


def parse_seconds(value, time_cell=False):
    if value is None:
        return None

    if isinstance(value, datetime.timedelta):
        return int(round(value.total_seconds()))

    if isinstance(value, (datetime.time, datetime.datetime)):
        # 无小时场景：Excel 时间 3:40 显示值按 3分40秒 处理（时->分，分->秒）
        return value.hour * 60 + value.minute

    if isinstance(value, (int, float)):
        if time_cell:
            # Excel 时间序列值（1 天 = 1.0），按 MM:SS 解释：serial*1440 = 秒
            return int(round(value * 1440))
        if value == int(value):
            return int(value)  # 纯整数按秒
        # 小数按 分.秒 解释（如 0.27 = 0分27秒 = 27 秒）
        sec = _parse_time_text(str(value))
        if sec is not None:
            if sec is _TIME_INVALID:
                return None
            return sec
        return int(round(value))

    if isinstance(value, str):
        s = _normalize_symbols(value).strip()
        if not s:
            return None
        # 过滤负数、科学计数法等奇怪符号
        if re.match(r"^-", s) or re.search(r"[eE]", s):
            return None
        sec = _parse_time_text(s)
        if sec is not None:
            if sec is _TIME_INVALID:
                return None
            return sec
        # 纯数字按秒（只允许纯整数或纯小数，不带奇怪后缀）
        if re.match(r"^\d+(\.\d+)?$", s):
            try:
                return int(round(float(s)))
            except ValueError:
                return None
        return None

    return None


def write_total(handle, kind, sheet_name, text, path):
    if kind == "xlsx":
        from openpyxl.styles import Font
        ws = handle[sheet_name]
        # 找已有的"总计"行（覆盖），否则追加到最后一个非空行下面
        total_row = None
        last = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v is not None and str(v).strip() != "":
                last = r
                if isinstance(v, str) and v.strip().startswith("总计"):
                    total_row = r
        if total_row is not None:
            cell = ws.cell(total_row, 1)
        else:
            cell = ws.cell(last + 1, 1)
        cell.value = text
        cell.font = Font(bold=True)
        handle.save(path)
        return True

    if kind == "csv":
        rows = handle.rows
        # 覆盖已有的总计行，否则追加
        replaced = False
        for i in range(len(rows) - 1, -1, -1):
            if rows[i] and rows[i][0].strip().startswith("总计"):
                rows[i] = [text]
                replaced = True
                break
        if not replaced:
            rows.append([text])
        with open(path, "w", encoding=handle.encoding, newline="") as f:
            csv.writer(f).writerows(rows)
        return True

    return False
